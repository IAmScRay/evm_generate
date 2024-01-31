from os import mkdir
from os.path import exists
from pathlib import Path

from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.workbook import Workbook
from web3 import Web3

INVERTED = "\033[7m"
RED = "\033[31m"
GREEN = "\033[32m"
RESET = "\033[0m"


def generate_wallets(amount: int):
    result = []

    for _ in range(amount):
        wallet = Web3().eth.account.create_with_mnemonic()

        data = {
            "address": str(wallet[0].address),
            "pk": wallet[0].key.hex(),
            "mnemonic": wallet[1]
        }
        result.append(data)

    return result


def main():

    amount = None
    while amount is None:
        try:
            a = int(input("Укажите желаемое кол-во кошельков: "))
        except ValueError:
            print(f"{INVERTED}Неверное значение!{RESET} {RED}Повторите попытку.{RESET}\n")
            continue

        amount = a

    try:
        mkdir(Path.cwd().joinpath("output"))
    except FileExistsError:
        pass

    filename = ""
    while filename == "":
        name = input("Укажите имя .xlsx-файла без расширения, куда будут записаны результаты: ")

        if exists(Path.cwd().joinpath("output").joinpath(f"{name}.xlsx")):
            print("Файл уже существует! Укажите уникальное имя.")
            continue
        else:
            filename = name

    wallets = generate_wallets(amount)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "EVM"

    A1 = sheet.cell(row=1, column=1)
    A1.value = "Адрес"

    B1 = sheet.cell(row=1, column=2)
    B1.value = "Приватный ключ"

    C1 = sheet.cell(row=1, column=3)
    C1.value = "Мнемоника"

    C1.font = B1.font = A1.font = Font(
        name="Arial",
        bold=True,
        size=16,
        color="FFFFFF"
    )

    C1.fill = B1.fill = A1.fill = PatternFill(
        fill_type="solid",
        start_color="000000",
        end_color="000000"
    )

    C1.alignment = B1.alignment = A1.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    row = 2
    address_width_set = False
    pk_width_set = False
    mnemonic_width_set = False
    for wallet in wallets:
        address_cell = sheet.cell(row=row, column=1)
        address_cell.value = wallet["address"]
        address_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        address_cell.font = Font(
            name="Consolas",
            bold=True,
            size=12
        )

        if not address_width_set:
            letter = address_cell.column_letter
            sheet.column_dimensions[letter].width = len(address_cell.value) * 1.5
            address_width_set = True

        pk_cell = sheet.cell(row=row, column=2)
        pk_cell.value = wallet["pk"]
        pk_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        pk_cell.font = Font(
            name="Consolas",
            bold=False,
            size=12
        )

        if not pk_width_set:
            letter = pk_cell.column_letter
            sheet.column_dimensions[letter].width = len(pk_cell.value) * 1.25
            pk_width_set = True

        mnemonic_cell = sheet.cell(row=row, column=3)
        mnemonic_cell.value = wallet["mnemonic"]
        mnemonic_cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )
        mnemonic_cell.font = Font(
            name="Consolas",
            bold=False,
            size=12
        )

        if not mnemonic_width_set:
            letter = mnemonic_cell.column_letter
            sheet.column_dimensions[letter].width = len(mnemonic_cell.value) * 1.5
            mnemonic_width_set = True

        row += 1

    workbook.save(f"output/{filename}.xlsx")
    print(INVERTED + GREEN + "Кошельки успешно сгенерированы!" + RESET)


if __name__ == "__main__":
    Web3().eth.account.enable_unaudited_hdwallet_features()

    main()
